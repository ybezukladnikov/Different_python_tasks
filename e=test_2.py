from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Border, Side
from datetime import datetime, timedelta

book = load_workbook('/Users/aroslavbezukladnikov/Desktop/excel_file/res1.xlsx')
ws = book.worksheets[0]
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 23
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 25
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 22
ws.column_dimensions["I"].width = 40
ws.column_dimensions["J"].width = 40

# Задаем стиль для ячейки
count = 2

while True:

    _cell = ws['G' + str(count)]
    now_date = datetime.now().date()
    try:
        date_time_obj = datetime.strptime(_cell.value, '%d.%m.%Y %H:%M')

    except TypeError:
        break

    if date_time_obj.date() < now_date:
        _cell.fill = PatternFill("solid", fgColor="00FF0000")

    elif date_time_obj.date() == now_date:
        _cell.fill = PatternFill("solid", fgColor="00FFCC00")

    else:
        _cell.fill = PatternFill("solid", fgColor="00FFFF00")

    count += 1

# print(date_time_obj.date())

# print(datetime.now().date())

# print(date_time_obj.date() == datetime.now().date())


# _cell.font = Font(size=10, underline='single', color='FFFFFF', bold=True, italic=True)

# Задаем цвет фона
# _cell.fill = PatternFill("solid", fgColor="00FFFF00")

# thins = Side(border_style="medium", color="FFC7CE")
# _cell.border = Border(top=thins, bottom=thins, left=thins, right=thins)

book.save('/Users/aroslavbezukladnikov/Desktop/excel_file/res1.xlsx')
